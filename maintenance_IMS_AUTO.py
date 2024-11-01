import json
import re
import logging
from tkinter import Tk, messagebox
from tkinter.filedialog import askopenfilename
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def get_excel_data(file_path):
    """
    Extract data from the Excel file and structure it into a list of dictionaries.
    """
    try:
        # Load the workbook and initialize the data list
        workbook = openpyxl.load_workbook(file_path)
    except InvalidFileException as e:
        logging.error(f"Invalid file: {file_path}. Error: {e}")
        return []
    except Exception as e:
        logging.error(f"Error loading workbook: {e}")
        return []

    data = []
    all_input_cells = []

    # Iterate through all sheets in the workbook
    for sheet in workbook.worksheets:
        sheet_data = {"sheet": sheet.title, "categories": []}
        current_category = None
        current_task = None
        input_columns = set()
        input_rows = set()
        comments_section = False

        # Identify columns with "Date inspected by who?" or "OK"
        for col in sheet.iter_cols(values_only=False):
            for cell in col:
                if cell.value is not None:
                    logging.debug(f"Checking cell {cell.coordinate} with value: {cell.value}")
                if cell.value in ["Date Inspected by who?", "OK", "OK?", "Date Inspected by who"]:
                    input_columns.add(openpyxl.utils.get_column_letter(cell.column))
        
        # Log identified input columns
        logging.info(f"Identified input columns in {sheet.title}: {sorted(input_columns)}")

        # Iterate through rows starting from the second row
        for row in sheet.iter_rows(min_row=2, values_only=False):
            cell_value = row[0].value
            if cell_value is None:
                continue

            cell_style = row[0].font
            cell_fill = row[0].fill  # Define cell_fill here

            # Check if the cell is a category (yellow background and bold text)
            if cell_fill.start_color and cell_fill.start_color.rgb == 'FFFFFF00' and cell_style.bold:
                current_category = cell_value
                sheet_data["categories"].append({"category": current_category, "tasks": []})
                current_task = None  # Reset current_task when a new category is found
                comments_section = False  # Reset comments_section when a new category is found
            # Check if the cell is a task (bold text)
            elif cell_style.bold and current_category and not comments_section:
                current_task = cell_value
                task_row = row[0].row  # Store the row of the current task
                input_rows.add(task_row)  # Collect input row coordinates
                logging.info(f"Identified task '{current_task}' in row {task_row}")
                logging.info(f"Task '{current_task}' identified in row {task_row}")
                sheet_data["categories"][-1]["tasks"].append({"task": current_task, "description": "", "inputs": {}})
            # Check if the cell is a description (non-bold text)
            elif current_task and not cell_style.bold and not comments_section:
                if sheet_data["categories"][-1]["tasks"][-1]["description"]:
                    sheet_data["categories"][-1]["tasks"][-1]["description"] += " " + cell_value
                else:
                    sheet_data["categories"][-1]["tasks"][-1]["description"] = cell_value
            # Check if the cell is part of the comments section
            elif cell_style.bold and current_category and comments_section:
                current_task = cell_value
                task_row = row[0].row  # Store the row of the current task
                input_rows.add(task_row)  # Collect input row coordinates
                logging.info(f"Identified comment '{current_task}' in row {task_row}")
                logging.info(f"Comment '{current_task}' identified in row {task_row}")
                sheet_data["categories"][-1]["tasks"].append({"task": current_task, "description": "", "inputs": {}})
            # Process input cells for the current task
            if current_category and current_task:
                for cell in row[1:sheet.max_column + 1]:  # Dynamically determine the range based on the actual number of columns
                    cell_coord = cell.coordinate
                    column_letter = openpyxl.utils.get_column_letter(cell.column)
                    if column_letter in input_columns and cell.row == task_row:
                        cell_value = cell.value
                        sheet_data["categories"][-1]["tasks"][-1]["inputs"][cell_coord] = cell_value if cell_value is not None else "no input"
                        logging.info(f"Processed input cell {cell_coord} with value: {cell_value}")
                        all_input_cells.append(cell_coord)  # Collect input cell coordinates

            # Check if the cell is the start of the comments section
            if cell_value == "Comments":
                comments_section = True
        
        # Sort the input rows
        logging.info(f"Identified input rows in {sheet.title}: {sorted(input_rows)}")
        # Append the sheet data to the main data list
        data.append(sheet_data)
    
    # Log all input cells
    logging.info(f"All input cells: {all_input_cells}")
    
    return data

def save_to_json(data, output_file):
    """
    Save the extracted data to a JSON file.
    """
    try:
        with open(output_file, 'w') as json_file:
            json.dump(data, json_file, indent=4)
    except Exception as e:
        logging.error(f"Error saving to JSON file: {e}")

def main(template=None):
    """
    Main function to run the script. Opens a file dialog for the user to select an Excel file,
    extracts the data, and saves it to a JSON file.
    """
    # Hide the root Tkinter window
    root = Tk()
    root.withdraw()
    
    # If template is provided, construct the path to the Excel file
    if template:
        file_path = os.path.join("IMS_TEMPLATE_COPIES", template, f"{template}.xlsx")
        if not os.path.exists(file_path):
            messagebox.showerror("Error", f"Template file not found: {file_path}")
            return
    else:
        # Open a file dialog for the user to select an Excel file
        file_path = askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])
        if not file_path:
            messagebox.showinfo("No file selected", "No file selected.")
            return
    
    # Extract data from the selected Excel file
    data = get_excel_data(file_path)
    
    if not data:
        messagebox.showerror("Error", "Failed to extract data from the Excel file.")
        return
    
    # Generate the output JSON file path
    base = file_path.rsplit('.', 1)[0]
    output_file = base + ".json"
    
    # Save the extracted data to the JSON file
    save_to_json(data, output_file)
    print(f"Data extracted and saved to {output_file}")

if __name__ == "__main__":
    main()
