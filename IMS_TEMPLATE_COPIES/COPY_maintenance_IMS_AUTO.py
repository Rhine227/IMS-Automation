import json
import re
from tkinter import Tk, messagebox
from tkinter.filedialog import askopenfilename
import openpyxl
import logging

def get_excel_data(file_path):
    """
    Extract data from the Excel file and structure it into a list of dictionaries.
    """
    workbook = openpyxl.load_workbook(file_path)
    data = []

    for sheet in workbook.worksheets:
        current_category = None
        current_task = None
        input_columns = set()
        INPUT_COLUMN_RANGE = range(1, 15)  # Columns B through O
        input_rows = set()

        # Identify columns with "Date inspected by who?" or "OK"
        for col in sheet.iter_cols(min_row=1, max_row=20, min_col=1, max_col=50, values_only=False):
            for cell in col:
                if cell.value is not None:
                    logging.debug(f"Checking cell {cell.coordinate} with value: {cell.value}")
                if cell.value in ["Date Inspected by who?", "OK"]:
                    input_columns.add(openpyxl.utils.get_column_letter(cell.column))

        for row in sheet.iter_rows(min_row=2, values_only=False):
            cell_value = row[0].value
            if cell_value is None:
                continue

            cell_style = row[0].font
            cell_font = row[0].fill  # Define cell_font here

            # Check if the cell is a category (yellow background and bold text)
            if cell_font.start_color.index == 'FFFFFF00' and cell_style.bold:
                current_category = cell_value
                data.append({"Category": current_category, "Tasks": []})
                current_task = None  # Reset current_task when a new category is found
            # Check if the cell is a task (bold text)
            elif cell_style.bold and current_category:
                current_task = cell_value
                data[-1]["Tasks"].append({"Task": current_task, "Description": "", "Inputs": {}})
                input_rows.add(cell.row)
            # Check if the cell is a description (non-bold text)
            elif current_task and not cell_style.bold:
                if data[-1]["Tasks"][-1]["Description"]:
                    data[-1]["Tasks"][-1]["Description"] += " " + cell_value
                else:
                    data[-1]["Tasks"][-1]["Description"] = cell_value

            # Process input cells for the current task
            if current_category and current_task:
                for cell in row[1:150]:  # Adjust the range to include columns B through O
                    cell_coord = cell.coordinate
                    column_letter = openpyxl.utils.get_column_letter(cell.column)
                    if column_letter in input_columns:
                        cell_value = cell.value
                        if cell_value is not None:
                            data[-1]["Tasks"][-1]["Inputs"][cell_coord] = cell_value
                        else:
                            data[-1]["Tasks"][-1]["Inputs"][cell_coord] = "no input"
                        logging.info(f"Processed cell {cell_coord} with value: {cell_value}")

        print(f"Identified input columns: {input_columns}")
        print(f"Identified input row: {input_rows}")
        input_cells = []
        for col in input_columns:
            for row in input_rows:
                cell_coord = f"{col}{row}"
                input_cells.append(cell_coord)
        
        input_cells.sort()
        print("Sorted input cells:", input_cells)
    
    return data

def save_to_json(data, output_file):
    """
    Save the extracted data to a JSON file.
    """
    with open(output_file, 'w') as json_file:
        json.dump(data, json_file, indent=4)

def main():
    """
    Main function to run the script. Opens a file dialog for the user to select an Excel file,
    extracts the data, and saves it to a JSON file.
    """
    Tk().withdraw()
    file_path = askopenfilename(title="Select Excel File", filetypes=[("Excel files", "*.xlsx *.xls")])
    if not file_path:
        messagebox.showinfo("No file selected", "No file selected.")
        return
    data = get_excel_data(file_path)
    base = file_path.rsplit('.', 1)[0]
    output_file = base + ".json"
    save_to_json(data, output_file)
    print(f"Data extracted and saved to {output_file}")

if __name__ == "__main__":
    main()
